# -*- coding: utf-8 -*-
"""
模板编辑器模块 - 积木式拖拽排列参数，自动添加单位（使用CSV列名）
"""

import os
import re
import tkinter as tk
from tkinter import ttk, messagebox, simpledialog

import pandas as pd
from openpyxl import load_workbook


class TemplateEditor:
    def __init__(self, parent, config_path, config_loader, template_name="标准模板", callback=None):
        self.parent = parent
        self.config_path = config_path
        self.config = config_loader
        self.template_name = template_name
        self.callback = callback
        
        self.selected_params = []
        self.available_params = []
        self.all_available_params = []
        
        self.drag_start_index = None
        self.drag_data = None
        
        self._build_window()
        self._load_available_params()
        self._load_current_template()
        self._update_ui()
    
    def _build_window(self):
        self.window = tk.Toplevel(self.parent)
        self.window.title(f"模板编辑器 - {self.template_name}")
        self.window.geometry("950x700")
        self.window.minsize(850, 600)
        self.window.transient(self.parent)
        
        main_frame = ttk.Frame(self.window, padding="10")
        main_frame.pack(fill=tk.BOTH, expand=True)
        
        top_frame = ttk.Frame(main_frame)
        top_frame.pack(fill=tk.X, pady=(0, 10))
        
        ttk.Label(top_frame, text=f"当前模板: {self.template_name}", font=('微软雅黑', 12, 'bold')).pack(side=tk.LEFT)
        
        btn_frame = ttk.Frame(top_frame)
        btn_frame.pack(side=tk.RIGHT)
        ttk.Button(btn_frame, text="💾 保存模板", command=self._save_template, width=12).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="📝 另存为新模板", command=self._save_as_template, width=14).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame, text="🗑 删除当前模板", command=self._delete_template, width=12).pack(side=tk.LEFT, padx=2)
        
        ttk.Separator(main_frame, orient='horizontal').pack(fill=tk.X, pady=5)
        
        body_frame = ttk.Frame(main_frame)
        body_frame.pack(fill=tk.BOTH, expand=True, pady=5)
        
        left_frame = ttk.LabelFrame(body_frame, text="📋 可用参数 (双击添加到右侧)", padding="5")
        left_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(0, 5))
        
        search_frame = ttk.Frame(left_frame)
        search_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(search_frame, text="🔍 筛选:").pack(side=tk.LEFT)
        self.search_var = tk.StringVar()
        self.search_var.trace('w', lambda *args: self._filter_available())
        search_entry = ttk.Entry(search_frame, textvariable=self.search_var)
        search_entry.pack(side=tk.LEFT, fill=tk.X, expand=True, padx=(5, 0))
        ttk.Button(search_frame, text="✕", command=lambda: self.search_var.set(""), width=3).pack(side=tk.LEFT, padx=(2, 0))
        
        list_frame = ttk.Frame(left_frame)
        list_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar_avail = ttk.Scrollbar(list_frame)
        scrollbar_avail.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.avail_listbox = tk.Listbox(
            list_frame,
            yscrollcommand=scrollbar_avail.set,
            font=('Consolas', 10),
            selectmode=tk.EXTENDED,
            activestyle='dotbox',
            exportselection=False
        )
        self.avail_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_avail.config(command=self.avail_listbox.yview)
        self.avail_listbox.bind('<Double-Button-1>', self._add_selected)
        
        self.avail_count_label = ttk.Label(left_frame, text="共 0 个可用参数 (已排除已选)")
        self.avail_count_label.pack(anchor=tk.W, pady=(3, 0))
        
        center_frame = ttk.Frame(body_frame, width=50)
        center_frame.pack(side=tk.LEFT, fill=tk.Y, padx=5)
        ttk.Label(center_frame, text="⟷", font=('Arial', 24)).pack(expand=True)
        ttk.Label(center_frame, text="双击添加", font=('微软雅黑', 9), foreground='gray').pack()
        ttk.Label(center_frame, text="拖拽排序", font=('微软雅黑', 9), foreground='gray').pack()
        
        right_frame = ttk.LabelFrame(body_frame, text="✅ 已选参数 (拖拽排序 · 双击移除)", padding="5")
        right_frame.pack(side=tk.LEFT, fill=tk.BOTH, expand=True, padx=(5, 0))
        
        hint_frame = ttk.Frame(right_frame)
        hint_frame.pack(fill=tk.X, pady=(0, 5))
        ttk.Label(hint_frame, text="按住参数拖动调整顺序 | 双击移除", font=('微软雅黑', 8), foreground='gray').pack(side=tk.LEFT)
        self.selected_count_label = ttk.Label(hint_frame, text="共 0 个参数", foreground='blue')
        self.selected_count_label.pack(side=tk.RIGHT)
        
        selected_frame = ttk.Frame(right_frame)
        selected_frame.pack(fill=tk.BOTH, expand=True)
        
        scrollbar_sel = ttk.Scrollbar(selected_frame)
        scrollbar_sel.pack(side=tk.RIGHT, fill=tk.Y)
        
        self.selected_listbox = tk.Listbox(
            selected_frame,
            yscrollcommand=scrollbar_sel.set,
            font=('Consolas', 10),
            selectmode=tk.SINGLE,
            activestyle='dotbox',
            exportselection=False,
            bg='#f0f8ff'
        )
        self.selected_listbox.pack(side=tk.LEFT, fill=tk.BOTH, expand=True)
        scrollbar_sel.config(command=self.selected_listbox.yview)
        
        self.selected_listbox.bind('<Double-Button-1>', self._remove_selected)
        self.selected_listbox.bind('<Button-1>', self._on_drag_start)
        self.selected_listbox.bind('<B1-Motion>', self._on_drag_motion)
        self.selected_listbox.bind('<ButtonRelease-1>', self._on_drag_end)
        
        preview_frame = ttk.LabelFrame(main_frame, text="📊 模板预览 (列顺序)", padding="5")
        preview_frame.pack(fill=tk.X, pady=(10, 0))
        
        self.preview_label = ttk.Label(preview_frame, text="(请添加参数)", font=('微软雅黑', 9), wraplength=850)
        self.preview_label.pack(fill=tk.X, pady=3)
        
        self.window.update_idletasks()
        x = self.parent.winfo_x() + (self.parent.winfo_width() - 950) // 2
        y = self.parent.winfo_y() + (self.parent.winfo_height() - 700) // 2
        self.window.geometry(f"+{x}+{y}")
    
    def _load_available_params(self):
        self.available_params = []
        self.all_available_params = []
        seen = set()
        
        for rule_id, rule in self.config.extract_rules.items():
            if rule_id not in seen:
                param_def = self.config.get_param_def(rule.get('param_id', ''))
                display_name = param_def.get('full_name', '') if param_def else ''
                if display_name:
                    label = f"{display_name} [{rule_id}]"
                else:
                    label = rule_id
                source = rule.get('source', '')
                item = (rule_id, label, source)
                self.available_params.append(item)
                self.all_available_params.append(item)
                seen.add(rule_id)
        
        meta_params = [
            ('DATE_META', '日期 [DATE_META]', 'META'),
            ('SAMPLE_ID_META', '样品ID [SAMPLE_ID_META]', 'META'),
            ('FEMUR_META', '股骨 [FEMUR_META]', 'META'),
        ]
        for meta_id, label, source in meta_params:
            if meta_id not in seen:
                item = (meta_id, label, source)
                self.available_params.append(item)
                self.all_available_params.append(item)
                seen.add(meta_id)
        
        self.available_params.sort(key=lambda x: x[1])
        self.all_available_params.sort(key=lambda x: x[1])
    
    def _load_current_template(self):
        self.selected_params = []
        columns = self.config.get_template_columns(self.template_name)
        for col in columns:
            self.selected_params.append(col['extract_id'])
    
    def _filter_available(self):
        keyword = self.search_var.get().strip().lower()
        self.avail_listbox.delete(0, tk.END)
        
        count = 0
        for rule_id, label, source in self.all_available_params:
            if rule_id in self.selected_params:
                continue
            if keyword:
                if keyword in label.lower() or keyword in rule_id.lower() or keyword in source.lower():
                    self.avail_listbox.insert(tk.END, label)
                    count += 1
            else:
                self.avail_listbox.insert(tk.END, label)
                count += 1
        
        self.avail_count_label.config(text=f"共 {count} 个可用参数 (已排除已选)")
    
    def _update_ui(self):
        self.selected_listbox.delete(0, tk.END)
        for rule_id in self.selected_params:
            label = rule_id
            for rid, display, _ in self.all_available_params:
                if rid == rule_id:
                    label = display
                    break
            self.selected_listbox.insert(tk.END, label)
        
        count = len(self.selected_params)
        self.selected_count_label.config(text=f"共 {count} 个参数")
        
        preview_text = " → ".join([f"{i+1}.{p}" for i, p in enumerate(self.selected_params)])
        if preview_text:
            if len(preview_text) > 200:
                preview_text = preview_text[:200] + "..."
            self.preview_label.config(text=preview_text)
        else:
            self.preview_label.config(text="(请添加参数)")
        
        self._filter_available()
    
    def _add_selected(self, event=None):
        selection = self.avail_listbox.curselection()
        if not selection:
            return
        
        label = self.avail_listbox.get(selection[0])
        rule_id = None
        for rid, display, _ in self.all_available_params:
            if display == label:
                rule_id = rid
                break
        
        if rule_id and rule_id not in self.selected_params:
            self.selected_params.append(rule_id)
            self._update_ui()
    
    def _remove_selected(self, event=None):
        selection = self.selected_listbox.curselection()
        if not selection:
            return
        
        idx = selection[0]
        if 0 <= idx < len(self.selected_params):
            self.selected_params.pop(idx)
            self._update_ui()
    
    def _on_drag_start(self, event):
        idx = self.selected_listbox.nearest(event.y)
        if idx < 0 or idx >= len(self.selected_params):
            return
        
        self.selected_listbox.selection_clear(0, tk.END)
        self.selected_listbox.selection_set(idx)
        self.selected_listbox.activate(idx)
        
        self.drag_start_index = idx
        self.drag_data = {
            'idx': idx,
            'item': self.selected_params[idx],
            'label': self.selected_listbox.get(idx)
        }
    
    def _on_drag_motion(self, event):
        if self.drag_data is None:
            return
        
        target_idx = self.selected_listbox.nearest(event.y)
        if target_idx < 0:
            target_idx = 0
        if target_idx >= len(self.selected_params):
            target_idx = len(self.selected_params) - 1
        
        if target_idx == self.drag_start_index:
            return
        
        item = self.drag_data['item']
        self.selected_params.pop(self.drag_start_index)
        self.selected_params.insert(target_idx, item)
        
        self._update_ui()
        
        self.selected_listbox.selection_clear(0, tk.END)
        self.selected_listbox.selection_set(target_idx)
        self.selected_listbox.activate(target_idx)
        
        self.drag_start_index = target_idx
    
    def _on_drag_end(self, event):
        self.drag_data = None
        self.drag_start_index = None
    
    def _collect_template_data(self):
        rows = []
        for idx, rule_id in enumerate(self.selected_params, start=1):
            rows.append({
                '模板名称': self.template_name,
                '列名': self._get_column_display_name(rule_id),
                '提取指令': rule_id,
                '顺序': idx
            })
        return rows
    
    # ★★★★★ 关键修改：使用 CSV列名（缩写）+ 单位 ★★★★★
    def _get_column_display_name(self, rule_id):
        # 元数据
        meta_names = {
            'DATE_META': '日期',
            'SAMPLE_ID_META': '样品ID',
            'FEMUR_META': '股骨'
        }
        if rule_id in meta_names:
            return meta_names[rule_id]

        # 从参数定义中获取 CSV列名 和 单位
        rule = self.config.get_extract_rule(rule_id)
        if rule:
            param_id = rule.get('param_id', '')
            param_def = self.config.get_param_def(param_id)
            if param_def:
                # ★ 优先使用 CSV列名（缩写）
                col_name = param_def.get('csv_column', '')
                unit = param_def.get('unit', '')
                if col_name:
                    if unit:
                        return f"{col_name} ({unit})"
                    return col_name
                # 如果 csv_column 为空，回退到完整名称
                full_name = param_def.get('full_name', '')
                if unit:
                    return f"{full_name} ({unit})"
                return full_name

        # 回退：从显示名中提取
        for rid, display, _ in self.all_available_params:
            if rid == rule_id:
                if ' [' in display:
                    return display.split(' [')[0]
                return display
        return rule_id
    
    def _save_template(self):
        if not self.selected_params:
            messagebox.showwarning("警告", "模板为空，请至少添加一个参数")
            return
        
        try:
            wb = load_workbook(self.config_path)
            
            if 'TemplateDef' not in wb.sheetnames:
                ws = wb.create_sheet('TemplateDef')
                ws.append(['模板名称', '列名', '提取指令', '顺序'])
            else:
                ws = wb['TemplateDef']
            
            rows_to_delete = []
            for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                if row and len(row) >= 1 and row[0].value == self.template_name:
                    rows_to_delete.append(idx)
            
            for idx in reversed(rows_to_delete):
                ws.delete_rows(idx)
            
            for idx, rule_id in enumerate(self.selected_params, start=1):
                col_name = self._get_column_display_name(rule_id)
                ws.append([self.template_name, col_name, rule_id, idx])
            
            wb.save(self.config_path)
            wb.close()
            
            self.config.load(self.config_path)
            
            messagebox.showinfo("成功", f"模板 '{self.template_name}' 已保存")
            
            if self.callback:
                self.callback()
            
        except Exception as e:
            messagebox.showerror("错误", f"保存失败: {e}")
    
    def _save_as_template(self):
        if not self.selected_params:
            messagebox.showwarning("警告", "模板为空，请至少添加一个参数")
            return
        
        new_name = simpledialog.askstring(
            "另存为",
            "请输入新模板名称:",
            parent=self.window,
            initialvalue=f"{self.template_name}_副本"
        )
        if not new_name:
            return
        
        if new_name in self.config.template_names:
            if not messagebox.askyesno("确认", f"模板 '{new_name}' 已存在，是否覆盖？"):
                return
        
        old_name = self.template_name
        self.template_name = new_name
        
        try:
            self._save_template()
            self.window.title(f"模板编辑器 - {self.template_name}")
            if self.callback:
                self.callback()
        except Exception as e:
            self.template_name = old_name
            messagebox.showerror("错误", f"保存失败: {e}")
    
    def _delete_template(self):
        if len(self.config.template_names) <= 1:
            messagebox.showwarning("警告", "至少保留一个模板")
            return
        
        if not messagebox.askyesno("确认删除", f"确定要删除模板 '{self.template_name}' 吗？\n此操作不可撤销！"):
            return
        
        try:
            wb = load_workbook(self.config_path)
            
            if 'TemplateDef' in wb.sheetnames:
                ws = wb['TemplateDef']
                
                rows_to_delete = []
                for idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
                    if row and len(row) >= 1 and row[0].value == self.template_name:
                        rows_to_delete.append(idx)
                
                for idx in reversed(rows_to_delete):
                    ws.delete_rows(idx)
                
                wb.save(self.config_path)
                wb.close()
                
                self.config.load(self.config_path)
                
                messagebox.showinfo("成功", f"模板 '{self.template_name}' 已删除")
                
                if self.callback:
                    self.callback()
                
                self.window.destroy()
            
        except Exception as e:
            messagebox.showerror("错误", f"删除失败: {e}")
    
    def destroy(self):
        self.window.destroy()
